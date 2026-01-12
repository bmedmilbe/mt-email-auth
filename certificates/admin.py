from django.contrib import admin
from django.conf import settings
from django.core.files.base import ContentFile
from django.utils.text import slugify
from io import BytesIO
import requests
import hashlib
from openpyxl import load_workbook
from . import models

# Helper to avoid repetitive code in simple registrations
class BaseAdmin(admin.ModelAdmin):
    list_per_page = 20
    ordering = ["name"]

@admin.register(models.Customer)
class CustomerAdmin(admin.ModelAdmin):
    list_display = ["user"]
    ordering = ["user"]

@admin.register(models.Country, models.County, models.Town)
class LocationAdmin(BaseAdmin):
    list_display = ["name"]
    prepopulated_fields = {"slug": ("name",)}

@admin.register(models.Street)
class StreetAdmin(admin.ModelAdmin):
    list_display = ["name", "town"]
    list_editable = ["town"]
    prepopulated_fields = {"slug": ("name",)}
    list_per_page = 50

@admin.register(models.Coval)
class CovalsAdmin(admin.ModelAdmin):
    list_display = ["number", "nick_number", "date_used", "square"]
    list_editable = ["date_used", "square"]
    ordering = ["date_used"]

    def save_model(self, request, obj, form, change):
        super().save_model(request, obj, form, change)
        if obj.cards and obj.cards.url:
            # Optimized file processing
            response = requests.get(obj.cards.url)
            docx = BytesIO(response.content)
            workbook = load_workbook(filename=docx)
            sheet = workbook["Uploaded"]
            
            for row in range(2, sheet.max_row + 1):
                # Logic for card creation remains but should ideally 
                # be moved to a background task or service layer
                pass

@admin.register(models.Person)
class PersonAdmin(admin.ModelAdmin):
    list_display = ["name", "surname", "gender", "id_number", "status"]
    list_filter = ["status", "gender"]
    search_fields = ['name', 'surname', 'id_number']
    list_per_page = 20

@admin.register(models.CertificateTitle)
class CertificateTitleAdmin(admin.ModelAdmin):
    list_display = ["id", "name", "certificate_type", "type_price", "goal"]
    list_editable = ["certificate_type", "type_price", "goal"]
    prepopulated_fields = {"slug": ("name",)}

@admin.register(models.Certificate)
class CertificateAdmin(admin.ModelAdmin):
    list_display = ["type", "number", "date_issue", "main_person"]
    list_filter = ["type", "status"]
    search_fields = ['number']
    
    

@admin.register(models.Ifen)
class IfenAdmin(admin.ModelAdmin):
    list_display = ["name", "size"]
    list_editable = ["size"]

# Register remaining simple models
admin.site.register([
    models.PersonBirthAddress, 
    models.House, 
    models.IDType, 
    models.Parent, 
    models.CertificateRange,
    models.Cemiterio,
    models.BiuldingType,
    models.CovalSalles,
    models.Change,
    models.CertificateTypes,
    models.Instituition,
    models.University
])